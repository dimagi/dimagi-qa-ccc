from datetime import datetime, timedelta
import json
import os
import time

from selenium.common import TimeoutException, NoSuchElementException
from selenium.webdriver import ActionChains, Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from utils.helpers import LocatorLoader, PROJECT_ROOT
from openpyxl import load_workbook

locators = LocatorLoader("locators/web_locators.yaml", platform="web")

class BaseWebPage:
    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(driver, 30)

    SUBMIT_BUTTON = locators.get("connect_opportunities_page", "submit_button")

    def find(self, locator):
        return self.driver.find_element(*locator)

    def find_all_elements(self, locator):
        return self.driver.find_elements(*locator)

    def click(self, locator):
        self.find(locator).click()

    def type(self, locator, text):
        el = self.find(locator)
        el.clear()
        el.send_keys(text)

    def wait_for_element(self, locator):
        return self.wait.until(EC.presence_of_element_located(locator))

    def is_selected(self, locator):
        try:
            element = self.driver.find_element(*locator)
            is_selected = element.is_selected()
        except TimeoutException:
            is_selected = False
        return bool(is_selected)

    def wait_for_clickable(self, locator):
        return self.wait.until(EC.element_to_be_clickable(locator))

    def click_element(self, locator):
        self.wait.until(EC.element_to_be_clickable(locator)).click()

    def type_element(self, locator, text):
        el = self.wait.until(EC.visibility_of_element_located(locator))
        el.clear()
        el.send_keys(text)

    def click_when_enabled(self, locator):
        self.wait.until(lambda d: d.find_element(*locator).is_enabled())
        self.driver.find_element(*locator).click()

    def get_text(self, locator):
        return self.wait.until(
            EC.visibility_of_element_located(locator)
        ).text

    def switch_to_latest_tab(self):
        self.driver.switch_to.window(self.driver.window_handles[-1])

    def switch_to_first_tab(self):
        self.driver.switch_to.window(self.driver.window_handles[0])

    def open_url_in_new_tab(self, url: str):
        self.driver.execute_script(f"window.open('{url}', '_blank');")
        self.switch_to_latest_tab()
        self.wait.until(EC.url_contains(url))
        assert url in self.driver.current_url, f"Expected URL '{url}' not opened. Found '{self.driver.current_url}'"

    def get_current_url(self):
        return self.driver.current_url

    def select_by_visible_text(self, dropdown_locator, text):
        element = self.wait.until(EC.presence_of_element_located(dropdown_locator))
        self.wait.until(lambda d: any(option.text.strip() == text for option in Select(element).options))
        Select(element).select_by_visible_text(text)

    def select_by_value(self, dropdown_locator, text):
        select_source = Select(self.driver.find_element(*dropdown_locator))
        select_source.select_by_value(text)

    def select_by_visible_text_manually(self, dropdown_locator, text):
        dropdown = self.wait.until(EC.element_to_be_clickable(dropdown_locator))

        # Click to focus
        dropdown.click()

        # Type partial text
        dropdown.send_keys(text)

        # Press Enter
        dropdown.send_keys(Keys.ENTER)

    def js_select_by_text(self, dropdown_locator, text):
        # Find element normally (supports any locator tuple)
        element = self.driver.find_element(*dropdown_locator)

        # Use JS to change value based on visible text
        self.driver.execute_script("""
            var select = arguments[0];
            var text = arguments[1];

            for (var i = 0; i < select.options.length; i++) {
                if (select.options[i].text.trim() === text.trim()) {
                    select.selectedIndex = i;
                    break;
                }
            }

            // Trigger native change event
            select.dispatchEvent(new Event('change', { bubbles: true }));

            // If HTMX exists, trigger via HTMX too
            if (window.htmx) {
                htmx.trigger(select, 'change');
            }
        """, element, text
                                   )

    def scroll_to_top(self):
        self.driver.execute_script("window.scrollTo(0, 0);")

    def scroll_to_bottom(self):
        self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

    def scroll_to_element(self, locator):
        element = self.driver.find_element(*locator)
        self.driver.execute_script("arguments[0].scrollIntoView();", element)

    def js_click(self, locator, timeout=10):
        element = WebDriverWait(self.driver, timeout, poll_frequency=0.25).until(
            EC.presence_of_element_located(locator),
            message=f"Couldn't find locator: {locator}"
            )
        self.driver.execute_script("arguments[0].click();", element)
        time.sleep(1)

    def scroll_into_view(self, locator, center=True):
        element = self.wait.until(EC.presence_of_element_located(locator))
        if center:
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
        else:
            self.driver.execute_script("arguments[0].scrollIntoView(true);",element)

    def scroll_until_element_visible(self, locator, max_scrolls=10):
        for _ in range(max_scrolls):
            try:
                element = self.find(*locator)
                if element.is_displayed():
                    return element
            except:
                pass
            self.driver.execute_script("window.scrollBy(arguments[0], arguments[1]);", 0, 500)
        raise TimeoutError("Element not visible after scrolling")

    def click_link_by_text(self, link_text: str):
        xpath = f"(//a[contains(.,'{link_text}')])[1]"
        element = self.wait.until(EC.presence_of_element_located((By.XPATH, xpath)))
        self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", element)
        self.scroll_to_element((By.XPATH, xpath))
        time.sleep(3)
        try:
            self.wait.until(EC.element_to_be_clickable((By.XPATH, xpath))).click()
        except:
            self.js_click((By.XPATH, xpath))
        self.wait_for_page_load()
        time.sleep(3)

    def is_breadcrumb_item_present(self, text: str) -> bool:
        xpath = f"//ul[contains(@class,'breadcrumb')]//a[contains(normalize-space(), '{text}')]"
        try:
            self.wait.until(EC.presence_of_element_located((By.XPATH, xpath)))
            return True
        except TimeoutException:
            return False

    def verify_text_in_url(self, expected_text: str, timeout: int = 10):
        WebDriverWait(self.driver, timeout).until(lambda driver: expected_text in driver.current_url)
        assert expected_text in self.driver.current_url, (
            f"Expected '{expected_text}' to be in URL, "
            f"but actual URL is '{self.driver.current_url}'"
        )

    def enter_date(self, locator, date_value: str):
        element = self.wait.until(EC.visibility_of_element_located(locator))
        try:
            element.clear()
            element.send_keys(date_value)
        except Exception:
            self.driver.execute_script(
                "arguments[0].value = arguments[1];"
                "arguments[0].dispatchEvent(new Event('change'));",
                element,
                date_value
            )
        actual_value = element.get_attribute("value")
        assert sorted(actual_value) == sorted(date_value), (
            f"Failed to set date. Expected '{date_value}', but got '{actual_value}'"
        )

    def click_submit_btn(self):
        self.scroll_into_view(self.SUBMIT_BUTTON)
        self.click_element(self.SUBMIT_BUTTON)

    def wait_for_page_load(self):
        self.wait.until(lambda d: d.execute_script("return document.readyState") == "complete")

    def navigate_backward(self):
        self.driver.back()
        self.wait_for_page_load()

    def navigate_forward(self):
        self.driver.forward()
        self.wait_for_page_load()

    def refresh_current_page(self):
        self.driver.refresh()
        self.wait_for_page_load()

    def find_element_or_fail(self, parent, by, locator, ele_name):
        try:
            return parent.find_element(by, locator)
        except NoSuchElementException:
            raise AssertionError(f"Element not found: {ele_name}")

    def write_payment_details_to_excel(self, params):
        project_root = PROJECT_ROOT
        file_path = os.path.join(project_root, "test_data", "make_payment.xlsx")
        wb = load_workbook(file_path)
        sheet = wb["Sheet1"]
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=2, column=col).value = None
        for index, value in enumerate(params, start=1):
            sheet.cell(row=2, column=index).value = value
        wb.save(file_path)
        print(params)

    def wait_for_page_to_load(self):
        self.wait.until(lambda driver: driver.execute_script("return document.readyState") == "complete")

    def reload_page(self):
        self.driver.refresh()

    def wait_for_js_alert_present(self):
        return self.wait.until(EC.alert_is_present())

    def xpath_literal(self, text):
        if "'" not in text:
            return f"'{text}'"
        if '"' not in text:
            return f'"{text}"'
        # If both quotes exist, use concat
        parts = text.split("'")
        return "concat(" + ", \"'\", ".join(f"'{p}'" for p in parts) + ")"

    def generate_date_range(self, days_to_add: int, opt=1):
        today = datetime.today()
        start = today + timedelta(days=2)
        end = start + timedelta(days=days_to_add)

        if opt==1:
            start_date = start.strftime("%m-%d-%Y")
            end_date = end.strftime("%m-%d-%Y")
        else:
            start_date = start.strftime("%d-%m-%Y")
            end_date = end.strftime("%d-%m-%Y")
        return start_date, end_date

    def is_displayed(self, locator):
        try:
            element = self.driver.find_element(*locator)
            is_displayed = element.is_displayed()
        except (TimeoutException, NoSuchElementException):
            is_displayed = False
        return bool(is_displayed)

    def is_present(self, locator):
        try:
            element = self.driver.find_element(*locator)
            is_displayed = True
        except NoSuchElementException:
            is_displayed = False
        return bool(is_displayed)

    def get_attribute(self, locator, attribute):
        element = self.driver.find_element(*locator)
        element_attribute = element.get_attribute(attribute)
        print(element_attribute)
        return element_attribute